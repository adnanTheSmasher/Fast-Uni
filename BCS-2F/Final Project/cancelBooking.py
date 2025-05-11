import sys
from openpyxl import load_workbook

room_to_cancel = sys.argv[1]
filename = "record.xlsx"

try:
    wb = load_workbook(filename)
    ws = wb.active
    found = False

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):  # skip header
        if str(row[0].value) == room_to_cancel:
            ws.delete_rows(i)
            found = True
            break

    if found:
        wb.save(filename)
        print("True")  # C++ checks for this
    else:
        print("False")

except FileNotFoundError:
    print("False")
