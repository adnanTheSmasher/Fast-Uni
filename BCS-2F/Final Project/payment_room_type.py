import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Missing CNIC argument")
    sys.exit(1)

cnic = sys.argv[1]
filename = "record.xlsx"
total_amount = 0

prices = {
    "Single": 500,
    "Double": 1500,
    "Suite": 5000
}

try:
    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True): 
        row_cnic = str(row[6]).strip()  
        if row_cnic == cnic:
            room_type = str(row[3]).strip().capitalize()  
            total_amount += prices.get(room_type, 0)

    print(total_amount)

except FileNotFoundError:
    print("0")
