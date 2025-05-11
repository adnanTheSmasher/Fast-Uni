import sys
import os
from openpyxl import load_workbook

room_to_check = sys.argv[1]
filename = "record.xlsx"

# If the file doesn't exist, treat all rooms as available
if not os.path.exists(filename):
    print("True")
    exit()

wb = load_workbook(filename)
ws = wb.active

# Check if the room number exists
for row in ws.iter_rows(min_row=2, values_only=True):
    if str(row[0]) == room_to_check:
        print("False")  # Room is already booked
        exit()

print("True")  # Room is available
