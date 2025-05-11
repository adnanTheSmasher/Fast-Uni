import json
from openpyxl import load_workbook
import os

with open("room_config.json") as f:
    all_rooms = json.load(f)

booked_rooms = set()
if os.path.exists("record.xlsx"):
    wb = load_workbook("record.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        booked_rooms.add(str(row[0]))

# Show only available rooms
available_rooms = {room: type_ for room, type_ in all_rooms.items() if room not in booked_rooms}

for room, type_ in available_rooms.items():
    print(f"Room {room}: {type_}")
