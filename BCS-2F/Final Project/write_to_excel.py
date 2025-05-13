import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import os

filename = "record.xlsx"

# Collect data
room_no = sys.argv[1]
room_no = int(room_no)
checkin = sys.argv[2]
checkout = sys.argv[3]
name = sys.argv[4]
mobile = sys.argv[5]
shanakti = sys.argv[6]
address = sys.argv[7]
paymentStatus = False
room_type = ""
amount = 0

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ye excel ko open or close karega data dalne k baad
if os.path.exists(filename):
    wb = load_workbook(filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # Add styled headers
    headers = ["Room No", "Check-in", "Check-out", "Type", "Name", "Mobile", "Shanakti", "Address", "Payment Status"]
    ws.append(headers)
    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.border = thin_border

if 1 <= room_no <= 3:
    room_type = "Single"

if 4 <= room_no <= 6:
    room_type = "Double"

if 8 <= room_no <= 10:
    room_type = "Suite"

# nai laine dalega
row_data = [room_no, checkin, checkout, room_type, name, mobile, shanakti, address, paymentStatus]
ws.append(row_data)

# nai line me border dalega
row_index = ws.max_row
for col in range(1, 9):  
    ws.cell(row=row_index, column=col).border = thin_border

# auto size feature
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# Save
wb.save(filename)
