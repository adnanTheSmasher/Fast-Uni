import sys
from openpyxl import load_workbook

room_to_find = sys.argv[1]
filename = "record.xlsx"

try:
    wb = load_workbook(filename)
    ws = wb.active
    found = False

    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        if str(row[0]) == room_to_find:
            print(f"Room Number: {row[0]}")
            print(f"Check-in Date: {row[1]}")
            print(f"Check-out Date: {row[2]}")
            print(f"Name: {row[3]}")
            print(f"Mobile: {row[4]}")
            print(f"Shanakti: {row[5]}")
            print(f"Address: {row[6]}")
            found = True
            break

    if not found:
        print("No room found")

except FileNotFoundError:
    print("No room found")
